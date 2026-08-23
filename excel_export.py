"""
Excel export for the screener pipeline. Writes the Stage 1 shortlist and
Stage 2 SEC filing catalysts into one workbook, each on its own sheet,
with basic readability formatting (bold headers, frozen header row,
auto-sized columns).
"""

import logging
import os

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from config import OUTPUT_DIR, SHORTLIST_FILENAME

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger(__name__)

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")


def _write_sheet(wb: Workbook, sheet_name: str, df: pd.DataFrame):
    ws = wb.create_sheet(title=sheet_name[:31])  # Excel sheet name cap

    if df.empty:
        ws.append(["No data"])
        return

    ws.append(list(df.columns))
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    ws.freeze_panes = "A2"

    for row in df.itertuples(index=False):
        ws.append(list(row))

    # auto-size columns (approximate, based on max content width per column)
    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = max(
            [len(str(col_name))] + [len(str(v)) for v in df[col_name].astype(str).tolist()]
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 45)


def export_to_excel(shortlist_df: pd.DataFrame, filings_df: pd.DataFrame | None = None,
                     scored_df: pd.DataFrame | None = None,
                     filename: str = SHORTLIST_FILENAME) -> str:
    """
    Writes the Stage 3 recommendations (if provided) as the primary first
    sheet, followed by the Stage 1 shortlist and Stage 2 filings detail
    sheets, to an .xlsx file in OUTPUT_DIR. Returns the full output path.
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    out_path = os.path.join(OUTPUT_DIR, filename)

    wb = Workbook()
    wb.remove(wb.active)  # drop the default blank sheet

    if scored_df is not None:
        _write_sheet(wb, "Recommendations", scored_df)

    _write_sheet(wb, "Stage1_Shortlist", shortlist_df)

    if filings_df is not None:
        _write_sheet(wb, "Stage2_SEC_Filings", filings_df)

    wb.save(out_path)
    log.info(f"Workbook written to {out_path} "
             f"({len(shortlist_df)} shortlist rows"
             + (f", {len(filings_df)} filing rows" if filings_df is not None else "")
             + (f", {len(scored_df)} scored rows" if scored_df is not None else "") + ").")
    return out_path
